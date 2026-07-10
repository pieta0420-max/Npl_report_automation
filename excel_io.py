"""DRM-aware Excel reader.

Tries openpyxl first (fast, no Excel required). If the file is wrapped in
DRM (IRM, DocumentSAFER, etc.) openpyxl can't even unzip it, so we fall back
to driving the user's own licensed Excel via COM automation. That works
because DRM decryption happens inside Excel using the user's own Windows
login/credentials -- the same access they'd have opening the file by hand.
"""
from __future__ import annotations

import zipfile
from pathlib import Path
from typing import Dict, List

SheetGrid = List[List[object]]


def _load_with_openpyxl(path: str) -> Dict[str, SheetGrid]:
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        sheets: Dict[str, SheetGrid] = {}
        for name in wb.sheetnames:
            ws = wb[name]
            max_row = ws.max_row or 0
            max_col = ws.max_column or 0
            grid: SheetGrid = []
            if max_row and max_col:
                for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col, values_only=True):
                    grid.append(list(row))
            sheets[name] = grid
        return sheets
    finally:
        wb.close()


def _load_with_com(path: str) -> Dict[str, SheetGrid]:
    import pythoncom
    import win32com.client as win32

    abs_path = str(Path(path).resolve())
    pythoncom.CoInitialize()
    # Dispatch (late-bound), not gencache.EnsureDispatch: the latter writes
    # generated wrapper modules to a cache directory that isn't reliably
    # writable/discoverable inside a frozen PyInstaller exe.
    excel = win32.Dispatch("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    try:
        wb = excel.Workbooks.Open(abs_path, ReadOnly=True)
        try:
            sheets: Dict[str, SheetGrid] = {}
            for sh in wb.Sheets:
                used = sh.UsedRange
                nrows = used.Rows.Count
                ncols = used.Columns.Count
                if nrows == 1 and ncols == 1:
                    values = [[used.Value]]
                else:
                    raw = used.Value  # COM returns a tuple of tuples
                    values = [list(row) for row in raw] if raw else []
                sheets[sh.Name] = values
            return sheets
        finally:
            wb.Close(False)
    finally:
        excel.Quit()
        pythoncom.CoUninitialize()


def is_drm_protected(path: str) -> bool:
    """Best-effort check: xlsx is a zip; DRM-wrapped files are not."""
    try:
        return not zipfile.is_zipfile(path)
    except OSError:
        return False


def load_workbook_any(path: str) -> Dict[str, SheetGrid]:
    """Load every sheet of an xlsx file as raw value grids.

    Returns dict[sheet_name] -> 2D list of cell values (row-major, 0-indexed).
    Falls back to COM automation transparently if the file is DRM-protected.
    """
    try:
        return _load_with_openpyxl(path)
    except zipfile.BadZipFile:
        return _load_with_com(path)
