"""Write the normalized sheets + 투심 summary tables to a fresh, DRM-free
.xlsx deliverable.

We build a brand-new workbook rather than writing into a copy of the
company's actual VF.xlsx template on purpose: that template carries
deal-specific formula cruft (stale Pool C/D blocks, hardcoded helper keys)
that isn't safe to inherit deal-to-deal (see schema.py / aggregator.py
docstrings). Colors, fonts, and the bilingual header layout below were
read directly off VF.xlsx via COM (Interior.Color / Font / NumberFormat)
so this looks like a native sibling of that file.

The 투심 summary sheet's count/OPB/채권총액/구성비 cells are written as live
Excel formulas (COUNTIFS/SUMIFS + cell-reference division for percentages),
not precomputed numbers -- they recalculate if a reviewer edits a row on
the data sheets. To make that possible without fragile array formulas, the
차주/담보/신용보증서/회생 sheets get a few extra "helper" columns (금액구간,
담보유형그룹, 지역그룹, 차주 OPB 등) holding the classification lookups /
cross-sheet joins that classification_config.json already drives; the
summary formulas then just COUNTIFS/SUMIFS against those columns.
"""
from __future__ import annotations

from dataclasses import dataclass
from typing import Dict, List, Optional, Sequence, Tuple

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from aggregator import AggregationResult
from classify import active_amount_buckets, amount_bucket_label, apply_property_alias, asset_type_label
from schema import CANONICAL_SCHEMAS, OUTPUT_SHEET_NAMES, ColumnDef, SheetSchema

# ---------------------------------------------------------------------------
# Styles (colors/fonts matched against WRB 2026-2 Program 투심보고서_VF.xlsx
# via win32com Interior.Color / Font / NumberFormat, read 2026-07-09)
# ---------------------------------------------------------------------------

FONT_FAMILY = "맑은 고딕"
FONT_SIZE = 10

PROGRAM_FONT = Font(name=FONT_FAMILY, bold=True, size=FONT_SIZE)
TITLE_FONT = Font(name=FONT_FAMILY, bold=True, size=FONT_SIZE)
SECTION_FONT = Font(name=FONT_FAMILY, bold=True, size=FONT_SIZE)
CATEGORY_FONT = Font(name=FONT_FAMILY, size=FONT_SIZE, color="FFFFFF")
HEADER_FONT = Font(name=FONT_FAMILY, size=FONT_SIZE)
HELPER_HEADER_FONT = Font(name=FONT_FAMILY, size=FONT_SIZE, italic=True)
DATA_FONT = Font(name=FONT_FAMILY, size=FONT_SIZE)
BOLD_DATA_FONT = Font(name=FONT_FAMILY, size=FONT_SIZE, bold=True)
NOTE_FONT = Font(name=FONT_FAMILY, italic=True, size=FONT_SIZE, color="666666")
POOL_FONT = Font(name=FONT_FAMILY, bold=True, size=FONT_SIZE, color="FFFFFF")
GROUP_LABEL_FONT = Font(name=FONT_FAMILY, size=FONT_SIZE, color="000000")

# 투심 tables start at column B, not A (leaves A free for the 담보유형/지역
# group-name annotation used by the property-type and region tables).
SUMMARY_COL_OFFSET = 1

CATEGORY_FILL = PatternFill("solid", fgColor="4472C4")   # data-sheet category/number rows
HEADER_FILL = PatternFill("solid", fgColor="BDD7EE")      # data-sheet EN/KR header rows
HELPER_FILL = PatternFill("solid", fgColor="FCE4D6")      # script-added helper columns (not from DD)
SUMMARY_HEADER_FILL = PatternFill("solid", fgColor="31849B")  # 투심 table header rows
POOL_FILL = PatternFill("solid", fgColor="31849B")

THIN = Side(style="thin", color="000000")
BOTTOM_BORDER = Border(bottom=THIN)
TOP_BORDER = Border(top=THIN)
TOP_BOTTOM_BORDER = Border(top=THIN, bottom=THIN)

CURRENCY_FMT = '_-* #,##0_-;-* #,##0_-;_-* "-"_-;_-@_-'
NUM_FMT = "#,##0"
PCT_FMT = "0.00%"

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")

NUMERIC_KEYS = {
    "opb_excl_prepaid", "prepaid_expense", "opb_incl_prepaid", "accrued_interest",
    "claim_total", "mortgage_amount_converted", "senior_mortgage_amount",
    "pre_cutoff_court_deposit", "appraisal_amount_total", "opb", "claim_amount",
    "initial_guarantee_amount", "guarantee_balance", "guarantee_balance_converted",
}


# ---------------------------------------------------------------------------
# Formula-string primitives
# ---------------------------------------------------------------------------

def _q(sheet_name: str) -> str:
    return f"'{sheet_name}'"


def _rng(sheet_name: str, col_letter: str) -> str:
    return f"{_q(sheet_name)}!${col_letter}:${col_letter}"


def _cell(sheet_name: Optional[str], col_letter: str, row: int) -> str:
    ref = f"${col_letter}${row}"
    return f"{_q(sheet_name)}!{ref}" if sheet_name else ref


def _lit(value) -> str:
    if isinstance(value, (int, float)):
        return str(value)
    return '"' + str(value).replace('"', '""') + '"'


def countifs_f(sheet: str, criteria: Sequence[Tuple[str, object]]) -> str:
    parts = [f"{_rng(sheet, col)},{_lit(val)}" for col, val in criteria]
    return "=COUNTIFS(" + ",".join(parts) + ")"


def sumifs_million_f(sheet: str, sum_col: str, criteria: Sequence[Tuple[str, object]]) -> str:
    parts = [_rng(sheet, sum_col)] + [f"{_rng(sheet, col)},{_lit(val)}" for col, val in criteria]
    return "=ROUND(SUMIFS(" + ",".join(parts) + ")/1000000,0)"


def sum_range_f(col_letter: str, first_row: int, last_row: int) -> str:
    return f"=SUM(${col_letter}${first_row}:${col_letter}${last_row})"


def div_f(numer_col: str, numer_row: int, denom_col: str, denom_row: int) -> str:
    return f"=IFERROR(${numer_col}${numer_row}/${denom_col}${denom_row},0)"


def SC(logical_col: int) -> int:
    """A 투심 table's logical 1-based column (1=its leftmost column) -> the
    actual worksheet column index, shifted so tables start at B."""
    return logical_col + SUMMARY_COL_OFFSET


def SL(logical_col: int) -> str:
    """Same as SC, but as a column letter (for formula range strings)."""
    return get_column_letter(SC(logical_col))


# ---------------------------------------------------------------------------
# Column map: canonical/helper key -> column letter, for a written sheet
# ---------------------------------------------------------------------------

ColumnMap = Dict[str, str]


@dataclass
class SheetContext:
    borrower_sheet: str
    borrower_cols: ColumnMap
    collateral_sheet: str
    collateral_cols: ColumnMap
    guarantee_sheet: str
    guarantee_cols: ColumnMap
    rehab_sheet: str
    rehab_cols: ColumnMap


def _write_title_block(ws: Worksheet, schema: SheetSchema, program_name: str) -> None:
    ws.cell(row=2, column=1, value=program_name).font = PROGRAM_FONT
    ws.cell(row=3, column=1, value=schema.header_text()).font = TITLE_FONT


def _write_category_and_number_rows(ws: Worksheet, headers: List[Tuple[str, str]]) -> None:
    """headers: list of (category, ignored) for columns 1..N; merges consecutive
    same-category cells on row 5 and fills row 5+6 with the VF blue."""
    n = len(headers)
    c = 1
    while c <= n:
        cat = headers[c - 1][0]
        span = 1
        while c + span - 1 < n and headers[c + span - 1][0] == cat:
            span += 1
        cell = ws.cell(row=5, column=c, value=cat)
        cell.font = CATEGORY_FONT
        cell.fill = CATEGORY_FILL
        cell.alignment = CENTER
        cell.border = BOTTOM_BORDER
        if span > 1:
            ws.merge_cells(start_row=5, start_column=c, end_row=5, end_column=c + span - 1)
        c += span
    for i in range(1, n + 1):
        cell = ws.cell(row=6, column=i, value=i)
        cell.font = CATEGORY_FONT
        cell.fill = CATEGORY_FILL
        cell.alignment = CENTER
        cell.border = BOTTOM_BORDER


def _write_bilingual_header(ws: Worksheet, col: int, header_en: str, header_kr: str, is_helper: bool = False) -> None:
    fill = HELPER_FILL if is_helper else HEADER_FILL
    font = HELPER_HEADER_FONT if is_helper else HEADER_FONT
    for r, text in ((7, header_en), (8, header_kr)):
        cell = ws.cell(row=r, column=col, value=text)
        cell.font = font
        cell.fill = fill
        cell.alignment = CENTER
        cell.border = BOTTOM_BORDER


def _write_canonical_columns(ws: Worksheet, schema: SheetSchema, df: pd.DataFrame,
                              skip_keys: Sequence[str] = ()) -> Tuple[ColumnMap, int]:
    """Writes every non-derived, non-skipped canonical column as plain values.
    Returns the column map and the next free column index (1-based)."""
    cols = [c for c in schema.columns if not c.derived and c.key not in skip_keys]
    _write_category_and_number_rows(ws, [(c.category, c.key) for c in cols])
    col_map: ColumnMap = {}
    for i, col in enumerate(cols, start=1):
        _write_bilingual_header(ws, i, col.header_en, col.header_kr)
        col_map[col.key] = get_column_letter(i)

    r = 9
    for _, row in df.iterrows():
        for i, col in enumerate(cols, start=1):
            val = row.get(col.key)
            if isinstance(val, float) and pd.isna(val):
                val = None
            cell = ws.cell(row=r, column=i, value=val)
            cell.font = DATA_FONT
            cell.border = BOTTOM_BORDER
            if col.key in NUMERIC_KEYS and val is not None:
                cell.number_format = CURRENCY_FMT
                cell.alignment = RIGHT
        r += 1
    return col_map, len(cols) + 1


def _write_helper_column(ws: Worksheet, col_idx: int, header_kr: str, header_en: str,
                          values: Optional[List] = None, formulas: Optional[List[str]] = None,
                          numfmt: Optional[str] = None) -> str:
    _write_bilingual_header(ws, col_idx, header_en, header_kr, is_helper=True)
    ws.cell(row=5, column=col_idx, value="자동 계산 (helper)").font = CATEGORY_FONT
    ws.cell(row=5, column=col_idx).fill = HELPER_FILL
    ws.cell(row=5, column=col_idx).alignment = CENTER
    ws.cell(row=6, column=col_idx, value=col_idx).font = CATEGORY_FONT
    ws.cell(row=6, column=col_idx).fill = HELPER_FILL

    source = formulas if formulas is not None else values
    r = 9
    for val in source:
        cell = ws.cell(row=r, column=col_idx, value=val)
        cell.font = DATA_FONT
        cell.border = BOTTOM_BORDER
        if numfmt:
            cell.number_format = numfmt
            cell.alignment = RIGHT
        r += 1
    return get_column_letter(col_idx)


def _write_borrower_key_column(ws: Worksheet, col_idx: int, col_map: ColumnMap, n_rows: int) -> str:
    """borrower_control_no is only unique WITHIN a pool -- once multiple
    Data Disk files (one per pool) are merged into one report, the same
    control number can be reused by an unrelated borrower in a different
    pool. Every VLOOKUP/SUMIFS/COUNTIFS that matches borrowers across sheets
    keys on this composite (pool, control_no) string instead of the raw
    control number, so a Pool B borrower can never match a Pool A row that
    happens to share the same number."""
    pool_col = col_map["pool_type"]
    key_col = col_map["borrower_control_no"]
    formulas = [f"={pool_col}{r}&\"|\"&{key_col}{r}" for r in range(9, 9 + n_rows)]
    return _write_helper_column(ws, col_idx, "차주키(자동계산)", "Borrower Key (auto)", formulas=formulas)


def _finalize_sheet(ws: Worksheet, n_cols: int, n_rows: int) -> None:
    for c in range(1, n_cols + 1):
        ws.column_dimensions[get_column_letter(c)].width = 14
    ws.freeze_panes = "A9"


# ---------------------------------------------------------------------------
# Sheet writers
# ---------------------------------------------------------------------------

def write_borrower_sheet(wb: Workbook, df: pd.DataFrame, config: dict, program_name: str) -> Tuple[str, ColumnMap]:
    schema = CANONICAL_SCHEMAS["borrower"]
    sheet_name = OUTPUT_SHEET_NAMES["borrower"]
    ws = wb.create_sheet(sheet_name)
    _write_title_block(ws, schema, program_name)

    col_map, next_col = _write_canonical_columns(ws, schema, df)

    col_map["borrower_key"] = _write_borrower_key_column(ws, next_col, col_map, len(df))
    next_col += 1

    bucket_thresholds, bucket_labels = active_amount_buckets(df, config)
    bucket_values = [amount_bucket_label(v, bucket_thresholds, bucket_labels) for v in df["opb_incl_prepaid"]]
    col_map["amount_bucket"] = _write_helper_column(
        ws, next_col, "금액구간 (자동계산)", "Amount Bucket (auto)", values=bucket_values)

    _finalize_sheet(ws, next_col, len(df))
    return sheet_name, col_map


def write_collateral_sheet(wb: Workbook, df: pd.DataFrame, config: dict,
                            unclassified_p: set, unclassified_r: set,
                            borrower_sheet: str, borrower_cols: ColumnMap,
                            program_name: str) -> Tuple[str, ColumnMap]:
    from classify import group_lookup

    schema = CANONICAL_SCHEMAS["collateral"]
    sheet_name = OUTPUT_SHEET_NAMES["collateral"]
    ws = wb.create_sheet(sheet_name)
    _write_title_block(ws, schema, program_name)

    col_map, next_col = _write_canonical_columns(ws, schema, df, skip_keys=("opb", "claim_amount"))

    col_map["borrower_key"] = _write_borrower_key_column(ws, next_col, col_map, len(df))
    next_col += 1

    # opb / claim_amount: proportional allocation as LIVE formulas (mirrors the
    # exact SUMIFS pattern found in VF.xlsx's own C-1 sheet formulas). Keyed
    # on the composite borrower_key (not the raw control number, which is
    # only unique within a pool) so a Pool B borrower's row can never match
    # a same-numbered Pool A borrower once multiple DDs are merged.
    key_col = col_map["borrower_key"]
    appraisal_col = col_map["appraisal_amount_total"]
    b_key_col = borrower_cols["borrower_key"]

    def _alloc_formula(source_col: str) -> List[str]:
        # Proportional split by appraisal amount; if a borrower's properties
        # have zero total appraisal (missing appraisal data), fall back to an
        # even split across that borrower's properties instead of /0.
        formulas = []
        for r in range(9, 9 + len(df)):
            numerator = f"SUMIFS({_rng(borrower_sheet, source_col)},{_rng(borrower_sheet, b_key_col)},${key_col}{r})"
            denom = f"SUMIFS({_rng(sheet_name, appraisal_col)},{_rng(sheet_name, key_col)},${key_col}{r})"
            count = f"COUNTIFS({_rng(sheet_name, key_col)},${key_col}{r})"
            proportional = f"{numerator}*${appraisal_col}{r}/{denom}"
            even_split = f"{numerator}/{count}"
            formulas.append(f"=ROUND(IFERROR({proportional},{even_split}),0)")
        return formulas

    opb_letter = get_column_letter(next_col)
    ws_col = next_col
    _write_bilingual_header(ws, ws_col, "OPB", "OPB")
    for r, f in zip(range(9, 9 + len(df)), _alloc_formula(borrower_cols["opb_incl_prepaid"])):
        cell = ws.cell(row=r, column=ws_col, value=f)
        cell.font = DATA_FONT
        cell.border = BOTTOM_BORDER
        cell.number_format = CURRENCY_FMT
        cell.alignment = RIGHT
    col_map["opb"] = opb_letter
    next_col += 1

    claim_letter = get_column_letter(next_col)
    _write_bilingual_header(ws, next_col, "Claim Amount", "채권액")
    for r, f in zip(range(9, 9 + len(df)), _alloc_formula(borrower_cols["claim_total"])):
        cell = ws.cell(row=r, column=next_col, value=f)
        cell.font = DATA_FONT
        cell.border = BOTTOM_BORDER
        cell.number_format = CURRENCY_FMT
        cell.alignment = RIGHT
    col_map["claim_amount"] = claim_letter
    next_col += 1

    # helper columns for the 투심 formulas: canonical property type + two group rollups
    canonical_types = [apply_property_alias(v, config) for v in df["property_type"]]
    col_map["property_type_canonical"] = _write_helper_column(
        ws, next_col, "담보유형(통합, 자동계산)", "Property Type (consolidated, auto)", values=canonical_types)
    next_col += 1

    p_groups = config["property_type_groups"]
    p_fallback = config["fallback_property_group"]
    prop_group_values = [group_lookup(v, p_groups, p_fallback, unclassified_p) for v in canonical_types]
    col_map["property_group"] = _write_helper_column(
        ws, next_col, "담보유형그룹(자동계산)", "Property Group (auto)", values=prop_group_values)
    next_col += 1

    r_groups = config["region_groups"]
    r_fallback = config["fallback_region_group"]
    region_group_values = [group_lookup(v, r_groups, r_fallback, unclassified_r) for v in df["property_addr1"]]
    col_map["region_group"] = _write_helper_column(
        ws, next_col, "지역그룹(자동계산)", "Region Group (auto)", values=region_group_values)
    next_col += 1

    _finalize_sheet(ws, next_col, len(df))
    return sheet_name, col_map


def _lookup_formula(key_cell: str, source_sheet: str, key_col: str, target_col: str) -> str:
    """INDEX/MATCH rather than VLOOKUP: VLOOKUP requires the target column to
    sit to the right of the key column within its search range, but the
    borrower_key helper column (the key used here) is appended after all of
    borrower sheet's canonical columns -- including opb/claim, which are
    common lookup targets -- so that ordering constraint doesn't hold."""
    key_rng = _rng(source_sheet, key_col)
    target_rng = _rng(source_sheet, target_col)
    return f"=IFERROR(INDEX({target_rng},MATCH({key_cell},{key_rng},0)),0)"


def write_guarantee_sheet(wb: Workbook, df: pd.DataFrame, borrower_sheet: str, borrower_cols: ColumnMap,
                           program_name: str) -> Tuple[str, ColumnMap]:
    schema = CANONICAL_SCHEMAS["guarantee"]
    sheet_name = OUTPUT_SHEET_NAMES["guarantee"]
    ws = wb.create_sheet(sheet_name)
    _write_title_block(ws, schema, program_name)

    col_map, next_col = _write_canonical_columns(ws, schema, df)

    col_map["borrower_key"] = _write_borrower_key_column(ws, next_col, col_map, len(df))
    next_col += 1
    key_col = col_map["borrower_key"]

    opb_formulas = [_lookup_formula(f"${key_col}{r}", borrower_sheet, borrower_cols["borrower_key"],
                                     borrower_cols["opb_incl_prepaid"]) for r in range(9, 9 + len(df))]
    col_map["borrower_opb"] = _write_helper_column(
        ws, next_col, "차주 OPB(자동계산)", "Borrower OPB (auto, INDEX/MATCH)", formulas=opb_formulas, numfmt=CURRENCY_FMT)
    next_col += 1

    # first-occurrence flag per borrower_key, so target_borrower_count /
    # target_opb don't double-count a borrower with >1 guarantee row
    first_formulas = [f"=IF(COUNTIF(${key_col}$9:${key_col}{r},${key_col}{r})=1,1,0)" for r in range(9, 9 + len(df))]
    col_map["is_first"] = _write_helper_column(
        ws, next_col, "최초등장(자동계산)", "First Occurrence (auto)", formulas=first_formulas)
    next_col += 1

    _finalize_sheet(ws, next_col, len(df))
    return sheet_name, col_map


def write_rehab_sheet(wb: Workbook, df: pd.DataFrame, borrower_sheet: str, borrower_cols: ColumnMap,
                       program_name: str) -> Tuple[str, ColumnMap]:
    schema = CANONICAL_SCHEMAS["rehab"]
    sheet_name = OUTPUT_SHEET_NAMES["rehab"]
    ws = wb.create_sheet(sheet_name)
    _write_title_block(ws, schema, program_name)

    col_map, next_col = _write_canonical_columns(ws, schema, df)

    col_map["borrower_key"] = _write_borrower_key_column(ws, next_col, col_map, len(df))
    next_col += 1
    key_col = col_map["borrower_key"]

    opb_formulas = [_lookup_formula(f"${key_col}{r}", borrower_sheet, borrower_cols["borrower_key"],
                                     borrower_cols["opb_incl_prepaid"]) for r in range(9, 9 + len(df))]
    col_map["borrower_opb"] = _write_helper_column(
        ws, next_col, "차주 OPB(자동계산)", "Borrower OPB (auto, INDEX/MATCH)", formulas=opb_formulas, numfmt=CURRENCY_FMT)
    next_col += 1

    claim_formulas = [_lookup_formula(f"${key_col}{r}", borrower_sheet, borrower_cols["borrower_key"],
                                       borrower_cols["claim_total"]) for r in range(9, 9 + len(df))]
    col_map["borrower_claim"] = _write_helper_column(
        ws, next_col, "차주 채권총액(자동계산)", "Borrower Claim Total (auto, INDEX/MATCH)",
        formulas=claim_formulas, numfmt=CURRENCY_FMT)
    next_col += 1

    _finalize_sheet(ws, next_col, len(df))
    return sheet_name, col_map


# ---------------------------------------------------------------------------
# 투심 summary sheet -- formula-driven
# ---------------------------------------------------------------------------

def _section_title(ws: Worksheet, row: int, text: str) -> int:
    ws.cell(row=row, column=SC(1), value=text).font = SECTION_FONT
    return row + 2


def _pool_header(ws: Worksheet, row: int, pool: str, n_cols: int, unit_note: str = "(단위: 백만원)") -> int:
    cell = ws.cell(row=row, column=SC(1), value=f"Pool {pool}")
    cell.font = POOL_FONT
    cell.fill = POOL_FILL
    ws.merge_cells(start_row=row, start_column=SC(1), end_row=row, end_column=SC(n_cols))
    row += 1
    ws.cell(row=row, column=SC(1), value=unit_note).font = NOTE_FONT
    return row + 1


def _table_header(ws: Worksheet, row: int, headers: List[str]) -> int:
    for i, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=SC(i), value=h)
        cell.font = CATEGORY_FONT
        cell.fill = SUMMARY_HEADER_FILL
        cell.alignment = CENTER
        cell.border = BOTTOM_BORDER
    return row + 1


def _apply_numfmt(ws: Worksheet, first_row: int, last_row: int, cols: Sequence[int], fmt: str = NUM_FMT) -> None:
    """cols are logical (1-based) 투심-table columns."""
    for r in range(first_row, last_row + 1):
        for c in cols:
            ws.cell(row=r, column=SC(c)).number_format = fmt


def _style_row(ws: Worksheet, row: int, n_cols: int, bold: bool = False) -> None:
    """bold=True marks a 소계/합계 row: top AND bottom borders box it off
    from the data rows above and whatever follows below. n_cols is the
    logical (1-based) column count of the table."""
    font = BOLD_DATA_FONT if bold else DATA_FONT
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=SC(c))
        cell.font = font
        if bold:
            cell.border = TOP_BOTTOM_BORDER


def _border_row(ws: Worksheet, row: int, n_cols: int, top: bool = False, bottom: bool = False,
                 include_col_a: bool = False) -> None:
    """Adds a border to an already-written row without touching its font
    (used for the group-boundary divider lines, rule 3)."""
    cols = range(0 if include_col_a else 1, n_cols + 1)
    for c in cols:
        cell = ws.cell(row=row, column=SC(c) if c else 1)
        existing = cell.border
        cell.border = Border(
            top=THIN if top else existing.top, bottom=THIN if bottom else existing.bottom,
            left=existing.left, right=existing.right,
        )


def _section_overall_summary(ws: Worksheet, row: int, borrower_df: pd.DataFrame,
                              agg_table: pd.DataFrame, config: dict, ctx: SheetContext) -> int:
    sheet = ctx.borrower_sheet
    pool_col, asset_col = ctx.borrower_cols["pool_type"], ctx.borrower_cols["asset_type"]
    opb_col, claim_col = ctx.borrower_cols["opb_incl_prepaid"], ctx.borrower_cols["claim_total"]
    label_to_raw = {v: k for k, v in config["asset_type_labels"].items()}
    headers = ["구분", "차주수", "구성비", "OPB", "구성비", "채권총액", "구성비"]

    pools = sorted(p for p in agg_table["pool_type"].unique() if p)
    subtotal_rows = []
    for pool in pools:
        pool_rows = agg_table[agg_table["pool_type"] == pool]
        row = _pool_header(ws, row, pool, len(headers))
        row = _table_header(ws, row, headers)
        first_data_row = row
        for _, r in pool_rows[pool_rows["row_kind"] == "data"].iterrows():
            raw_asset = label_to_raw.get(r["category"], r["category"])
            crit = [(pool_col, pool), (asset_col, raw_asset)]
            ws.cell(row=row, column=SC(1), value=r["category"])
            ws.cell(row=row, column=SC(2), value=countifs_f(sheet, crit))
            ws.cell(row=row, column=SC(4), value=sumifs_million_f(sheet, opb_col, crit))
            ws.cell(row=row, column=SC(6), value=sumifs_million_f(sheet, claim_col, crit))
            _style_row(ws, row, len(headers))
            row += 1
        last_data_row = row - 1
        subtotal_row = row
        ws.cell(row=row, column=SC(1), value="소계")
        ws.cell(row=row, column=SC(2), value=sum_range_f(SL(2), first_data_row, last_data_row))
        ws.cell(row=row, column=SC(4), value=sum_range_f(SL(4), first_data_row, last_data_row))
        ws.cell(row=row, column=SC(6), value=sum_range_f(SL(6), first_data_row, last_data_row))
        for r in range(first_data_row, subtotal_row + 1):
            ws.cell(row=r, column=SC(3), value=div_f(SL(2), r, SL(2), subtotal_row))
            ws.cell(row=r, column=SC(5), value=div_f(SL(4), r, SL(4), subtotal_row))
            ws.cell(row=r, column=SC(7), value=div_f(SL(6), r, SL(6), subtotal_row))
            for c in (3, 5, 7):
                ws.cell(row=r, column=SC(c)).number_format = PCT_FMT
        _apply_numfmt(ws, first_data_row, subtotal_row, (2, 4, 6))
        _style_row(ws, subtotal_row, len(headers), bold=True)
        subtotal_rows.append(subtotal_row)
        row = subtotal_row + 2

    if subtotal_rows:
        total_row = row
        ws.cell(row=row, column=SC(1), value="합계")
        ws.cell(row=row, column=SC(2), value="=" + "+".join(f"${SL(2)}${r}" for r in subtotal_rows))
        ws.cell(row=row, column=SC(4), value="=" + "+".join(f"${SL(4)}${r}" for r in subtotal_rows))
        ws.cell(row=row, column=SC(6), value="=" + "+".join(f"${SL(6)}${r}" for r in subtotal_rows))
        _apply_numfmt(ws, total_row, total_row, (2, 4, 6))
        _style_row(ws, total_row, len(headers), bold=True)
        row = total_row + 2
    return row


def _section_borrower_type(ws: Worksheet, row: int, agg_table: pd.DataFrame, ctx: SheetContext) -> int:
    sheet = ctx.borrower_sheet
    pool_col, type_col = ctx.borrower_cols["pool_type"], ctx.borrower_cols["borrower_type"]
    opb_col, claim_col = ctx.borrower_cols["opb_incl_prepaid"], ctx.borrower_cols["claim_total"]
    headers = ["구분", "차주수", "구성비", "OPB", "구성비", "채권총액", "구성비"]

    for pool in sorted(agg_table["pool_type"].unique()):
        pool_rows = agg_table[agg_table["pool_type"] == pool]
        row = _pool_header(ws, row, pool, len(headers))
        row = _table_header(ws, row, headers)
        first_data_row = row
        for _, r in pool_rows[pool_rows["row_kind"] == "data"].iterrows():
            crit = [(pool_col, pool), (type_col, r["category"])]
            ws.cell(row=row, column=SC(1), value=r["category"])
            ws.cell(row=row, column=SC(2), value=countifs_f(sheet, crit))
            ws.cell(row=row, column=SC(4), value=sumifs_million_f(sheet, opb_col, crit))
            ws.cell(row=row, column=SC(6), value=sumifs_million_f(sheet, claim_col, crit))
            _style_row(ws, row, len(headers))
            row += 1
        last_data_row = row - 1
        subtotal_row = row
        ws.cell(row=row, column=SC(1), value="소계")
        ws.cell(row=row, column=SC(2), value=sum_range_f(SL(2), first_data_row, last_data_row))
        ws.cell(row=row, column=SC(4), value=sum_range_f(SL(4), first_data_row, last_data_row))
        ws.cell(row=row, column=SC(6), value=sum_range_f(SL(6), first_data_row, last_data_row))
        for r in range(first_data_row, subtotal_row + 1):
            ws.cell(row=r, column=SC(3), value=div_f(SL(2), r, SL(2), subtotal_row))
            ws.cell(row=r, column=SC(5), value=div_f(SL(4), r, SL(4), subtotal_row))
            ws.cell(row=r, column=SC(7), value=div_f(SL(6), r, SL(6), subtotal_row))
            for c in (3, 5, 7):
                ws.cell(row=r, column=SC(c)).number_format = PCT_FMT
        _apply_numfmt(ws, first_data_row, subtotal_row, (2, 4, 6))
        _style_row(ws, subtotal_row, len(headers), bold=True)
        row = subtotal_row + 2
    return row


def _section_amount_bucket(ws: Worksheet, row: int, agg_table: pd.DataFrame, ctx: SheetContext) -> int:
    sheet = ctx.borrower_sheet
    pool_col, bucket_col = ctx.borrower_cols["pool_type"], ctx.borrower_cols["amount_bucket"]
    opb_col = ctx.borrower_cols["opb_incl_prepaid"]
    headers = ["구분", "차주수", "OPB", "구성비", "누적 차주수", "누적 OPB", "누적 구성비"]

    for pool in sorted(agg_table["pool_type"].unique()):
        pool_rows = agg_table[agg_table["pool_type"] == pool]
        row = _pool_header(ws, row, pool, len(headers))
        row = _table_header(ws, row, headers)
        first_data_row = row
        for _, r in pool_rows[pool_rows["row_kind"] == "data"].iterrows():
            crit = [(pool_col, pool), (bucket_col, r["category"])]
            ws.cell(row=row, column=SC(1), value=r["category"])
            ws.cell(row=row, column=SC(2), value=countifs_f(sheet, crit))
            ws.cell(row=row, column=SC(3), value=sumifs_million_f(sheet, opb_col, crit))
            _style_row(ws, row, len(headers))
            row += 1
        last_data_row = row - 1
        subtotal_row = row
        ws.cell(row=row, column=SC(1), value="소계")
        ws.cell(row=row, column=SC(2), value=sum_range_f(SL(2), first_data_row, last_data_row))
        ws.cell(row=row, column=SC(3), value=sum_range_f(SL(3), first_data_row, last_data_row))
        for i, r in enumerate(range(first_data_row, last_data_row + 1)):
            ws.cell(row=r, column=SC(4), value=div_f(SL(3), r, SL(3), subtotal_row))
            ws.cell(row=r, column=SC(4)).number_format = PCT_FMT
            ws.cell(row=r, column=SC(5), value=sum_range_f(SL(2), first_data_row, r))
            ws.cell(row=r, column=SC(6), value=sum_range_f(SL(3), first_data_row, r))
            ws.cell(row=r, column=SC(7), value=div_f(SL(6), r, SL(3), subtotal_row))
            ws.cell(row=r, column=SC(7)).number_format = PCT_FMT
        _apply_numfmt(ws, first_data_row, subtotal_row, (2, 3))
        _apply_numfmt(ws, first_data_row, last_data_row, (5, 6))
        _style_row(ws, subtotal_row, len(headers), bold=True)
        row = subtotal_row + 2
    return row


def _section_group_table(ws: Worksheet, row: int, agg_table: pd.DataFrame, sheet: str,
                          pool_col: str, item_col: str, group_col: str, opb_col: str,
                          item_key: str, item_header: str) -> int:
    """Shared by 담보물 종류별/지역별 분류. Column H (유사유형 구성비) holds
    the group name as "(그룹명)" text on the row just below the row that
    carries that group's % value -- or, if a group only has one member, on
    an extra row inserted just to carry that label. A divider line separates
    each group's block from the next; the 합계 row also totals the
    group-rollup columns (물건수/OPB/구성비), not just the per-item ones."""
    headers = [item_header, "물건수", "OPB", "구성비", "유사유형 물건수", "유사유형 OPB", "유사유형 구성비"]

    for pool in sorted(agg_table["pool_type"].unique()):
        pool_rows = agg_table[agg_table["pool_type"] == pool]
        row = _pool_header(ws, row, pool, len(headers))
        row = _table_header(ws, row, headers)
        data = pool_rows[pool_rows["row_kind"] == "data"]

        blocks: List[Tuple[object, list]] = []
        for _, r in data.iterrows():
            if blocks and blocks[-1][0] == r["group"]:
                blocks[-1][1].append(r)
            else:
                blocks.append((r["group"], [r]))

        first_data_row = row
        pct_rows: List[int] = []
        group_first_rows: List[int] = []

        for block_idx, (group_name, items) in enumerate(blocks):
            block_start_row = row
            for i, r in enumerate(items):
                crit = [(pool_col, pool), (item_col, r[item_key])]
                ws.cell(row=row, column=SC(1), value=r[item_key])
                ws.cell(row=row, column=SC(2), value=countifs_f(sheet, crit))
                ws.cell(row=row, column=SC(3), value=sumifs_million_f(sheet, opb_col, crit))
                if i == 0:
                    gcrit = [(pool_col, pool), (group_col, group_name)]
                    ws.cell(row=row, column=SC(5), value=countifs_f(sheet, gcrit))
                    ws.cell(row=row, column=SC(6), value=sumifs_million_f(sheet, opb_col, gcrit))
                    group_first_rows.append(row)
                _style_row(ws, row, len(headers))
                pct_rows.append(row)
                row += 1

            if len(items) >= 2:
                label_row = block_start_row + 1
                block_last_row = row - 1
            else:
                label_row = row
                _style_row(ws, label_row, len(headers))
                block_last_row = label_row
                row += 1
            label_cell = ws.cell(row=label_row, column=SC(7), value=f"({group_name})")
            label_cell.font = GROUP_LABEL_FONT
            label_cell.alignment = RIGHT

            if block_idx < len(blocks) - 1:
                _border_row(ws, block_last_row, len(headers), bottom=True)

        last_data_row = row - 1
        total_row = row
        ws.cell(row=row, column=SC(1), value="합계")
        ws.cell(row=row, column=SC(2), value=sum_range_f(SL(2), first_data_row, last_data_row))
        ws.cell(row=row, column=SC(3), value=sum_range_f(SL(3), first_data_row, last_data_row))
        ws.cell(row=row, column=SC(5), value=sum_range_f(SL(5), first_data_row, last_data_row))
        ws.cell(row=row, column=SC(6), value=sum_range_f(SL(6), first_data_row, last_data_row))
        for r in pct_rows:
            ws.cell(row=r, column=SC(4), value=div_f(SL(3), r, SL(3), total_row))
            ws.cell(row=r, column=SC(4)).number_format = PCT_FMT
        for r in group_first_rows:
            ws.cell(row=r, column=SC(7), value=div_f(SL(6), r, SL(3), total_row))
            ws.cell(row=r, column=SC(7)).number_format = PCT_FMT
        ws.cell(row=total_row, column=SC(7), value=sum_range_f(SL(7), first_data_row, last_data_row))
        ws.cell(row=total_row, column=SC(7)).number_format = PCT_FMT
        _apply_numfmt(ws, first_data_row, total_row, (2, 3, 5, 6))
        _style_row(ws, total_row, len(headers), bold=True)
        row = total_row + 2
    return row


def _section_property_type(ws: Worksheet, row: int, agg_table: pd.DataFrame, ctx: SheetContext) -> int:
    return _section_group_table(
        ws, row, agg_table, ctx.collateral_sheet,
        ctx.collateral_cols["pool_type"], ctx.collateral_cols["property_type_canonical"],
        ctx.collateral_cols["property_group"], ctx.collateral_cols["opb"],
        "property_type", "담보물 종류")


def _section_region(ws: Worksheet, row: int, agg_table: pd.DataFrame, ctx: SheetContext) -> int:
    return _section_group_table(
        ws, row, agg_table, ctx.collateral_sheet,
        ctx.collateral_cols["pool_type"], ctx.collateral_cols["property_addr1"],
        ctx.collateral_cols["region_group"], ctx.collateral_cols["opb"],
        "region", "담보 소재지")


def _section_matrix(ws: Worksheet, row: int, matrix_df: pd.DataFrame, ctx: SheetContext) -> int:
    """(*) 지역별/물건별 OPB 비중 -- kept as computed values (a % of a % of a
    ratio isn't worth a fragile nested SUMPRODUCT formula); still fully
    derived from the same helper columns as the tables above.

    Each pool's own columns are ordered by that pool's OWN composition
    ratio (summed across every region row), descending -- not one shared
    order across all pools, since which property type dominates can
    differ pool to pool."""
    if matrix_df.empty:
        return row
    all_prop_groups = [c for c in matrix_df.columns if c not in ("pool_type", "region_group")]
    for pool in matrix_df["pool_type"].unique():
        pool_df = matrix_df[matrix_df["pool_type"] == pool]
        prop_groups = sorted(all_prop_groups, key=lambda pg: -pool_df[pg].sum())
        cell = ws.cell(row=row, column=SC(1), value=f"Pool {pool} (*) 지역별/물건별 OPB 비중")
        cell.font = POOL_FONT
        cell.fill = POOL_FILL
        ws.merge_cells(start_row=row, start_column=SC(1), end_row=row, end_column=SC(len(prop_groups) + 1))
        row += 1
        row = _table_header(ws, row, ["구분"] + prop_groups)
        for _, data_row in pool_df.iterrows():
            ws.cell(row=row, column=SC(1), value=data_row["region_group"]).font = DATA_FONT
            for i, pg in enumerate(prop_groups, start=2):
                cell = ws.cell(row=row, column=SC(i), value=data_row[pg])
                cell.number_format = PCT_FMT
                cell.font = DATA_FONT
            row += 1
        row += 1
    return row + 1


def _section_foreclosure(ws: Worksheet, row: int, agg_table: pd.DataFrame, ctx: SheetContext) -> int:
    sheet = ctx.collateral_sheet
    pool_col, status_col = ctx.collateral_cols["pool_type"], ctx.collateral_cols["foreclosure_status"]
    opb_col, claim_col = ctx.collateral_cols["opb"], ctx.collateral_cols["claim_amount"]
    label_to_raw = {"경매 개시": "Filed", "경매 미개시": "Not Filed"}
    headers = ["구분", "담보물건수", "OPB", "구성비", "채권총액", "구성비"]

    for pool in sorted(agg_table["pool_type"].unique()):
        pool_rows = agg_table[agg_table["pool_type"] == pool]
        row = _pool_header(ws, row, pool, len(headers))
        row = _table_header(ws, row, headers)
        first_data_row = row
        for _, r in pool_rows[pool_rows["row_kind"] == "data"].iterrows():
            raw = label_to_raw.get(r["category"], r["category"])
            crit = [(pool_col, pool), (status_col, raw)]
            ws.cell(row=row, column=SC(1), value=r["category"])
            ws.cell(row=row, column=SC(2), value=countifs_f(sheet, crit))
            ws.cell(row=row, column=SC(3), value=sumifs_million_f(sheet, opb_col, crit))
            ws.cell(row=row, column=SC(5), value=sumifs_million_f(sheet, claim_col, crit))
            _style_row(ws, row, len(headers))
            row += 1
        last_data_row = row - 1
        total_row = row
        ws.cell(row=row, column=SC(1), value="합계")
        ws.cell(row=row, column=SC(2), value=sum_range_f(SL(2), first_data_row, last_data_row))
        ws.cell(row=row, column=SC(3), value=sum_range_f(SL(3), first_data_row, last_data_row))
        ws.cell(row=row, column=SC(5), value=sum_range_f(SL(5), first_data_row, last_data_row))
        for r in range(first_data_row, total_row + 1):
            ws.cell(row=r, column=SC(4), value=div_f(SL(3), r, SL(3), total_row))
            ws.cell(row=r, column=SC(6), value=div_f(SL(5), r, SL(5), total_row))
            for c in (4, 6):
                ws.cell(row=r, column=SC(c)).number_format = PCT_FMT
        _apply_numfmt(ws, first_data_row, total_row, (2, 3, 5))
        _style_row(ws, total_row, len(headers), bold=True)
        row = total_row + 2
    return row


def _section_rehab(ws: Worksheet, row: int, agg_table: pd.DataFrame, ctx: SheetContext) -> int:
    if agg_table.empty:
        ws.cell(row=row, column=SC(1), value="(해당 데이터 없음)").font = NOTE_FONT
        return row + 2
    sheet = ctx.rehab_sheet
    pool_col, state_col = ctx.rehab_cols["pool_type"], ctx.rehab_cols["state"]
    opb_col, claim_col = ctx.rehab_cols["borrower_opb"], ctx.rehab_cols["borrower_claim"]
    headers = ["구분", "차주수", "OPB", "구성비", "채권총액", "구성비"]

    for pool in sorted(agg_table["pool_type"].unique()):
        pool_rows = agg_table[agg_table["pool_type"] == pool]
        row = _pool_header(ws, row, pool, len(headers))
        row = _table_header(ws, row, headers)
        first_data_row = row
        for _, r in pool_rows[pool_rows["row_kind"] == "data"].iterrows():
            crit = [(pool_col, pool), (state_col, r["category"])]
            ws.cell(row=row, column=SC(1), value=r["category"])
            ws.cell(row=row, column=SC(2), value=countifs_f(sheet, crit))
            ws.cell(row=row, column=SC(3), value=sumifs_million_f(sheet, opb_col, crit))
            ws.cell(row=row, column=SC(5), value=sumifs_million_f(sheet, claim_col, crit))
            _style_row(ws, row, len(headers))
            row += 1
        last_data_row = row - 1
        total_row = row
        ws.cell(row=row, column=SC(1), value="합계")
        ws.cell(row=row, column=SC(2), value=sum_range_f(SL(2), first_data_row, last_data_row))
        ws.cell(row=row, column=SC(3), value=sum_range_f(SL(3), first_data_row, last_data_row))
        ws.cell(row=row, column=SC(5), value=sum_range_f(SL(5), first_data_row, last_data_row))
        for r in range(first_data_row, total_row + 1):
            ws.cell(row=r, column=SC(4), value=div_f(SL(3), r, SL(3), total_row))
            ws.cell(row=r, column=SC(6), value=div_f(SL(5), r, SL(5), total_row))
            for c in (4, 6):
                ws.cell(row=r, column=SC(c)).number_format = PCT_FMT
        _apply_numfmt(ws, first_data_row, total_row, (2, 3, 5))
        _style_row(ws, total_row, len(headers), bold=True)
        row = total_row + 2
    return row


def _section_guarantee(ws: Worksheet, row: int, pools: List[str], ctx: SheetContext) -> int:
    b_pool_col, b_opb_col = ctx.borrower_cols["pool_type"], ctx.borrower_cols["opb_incl_prepaid"]
    g_pool_col = ctx.guarantee_cols["pool_type"]
    g_opb_col, g_is_first = ctx.guarantee_cols["borrower_opb"], ctx.guarantee_cols["is_first"]
    g_balance_col = ctx.guarantee_cols["guarantee_balance_converted"]
    g_guarantor_col = ctx.guarantee_cols["guarantor"]
    headers = ["Pool", "전체 OPB", "대상 차주수", "대상 OPB", "신용보증서 유효잔액", "전체OPB 대비 유효잔액비율"]

    row = _table_header(ws, row, headers)
    for pool in pools:
        # 신용보증서 유효잔액: excludes MCI-classified guarantees (보증기관명에
        # "MCI"가 포함된 행은 제외) -- confirmed with user 2026-07-10.
        balance_criteria = [(g_pool_col, pool), (g_guarantor_col, "<>*MCI*")]
        ws.cell(row=row, column=SC(1), value=pool)
        ws.cell(row=row, column=SC(2), value=sumifs_million_f(ctx.borrower_sheet, b_opb_col, [(b_pool_col, pool)]))
        ws.cell(row=row, column=SC(3), value=countifs_f(ctx.guarantee_sheet, [(g_pool_col, pool), (g_is_first, 1)]))
        ws.cell(row=row, column=SC(4), value=sumifs_million_f(
            ctx.guarantee_sheet, g_opb_col, [(g_pool_col, pool), (g_is_first, 1)]))
        ws.cell(row=row, column=SC(5), value=sumifs_million_f(ctx.guarantee_sheet, g_balance_col, balance_criteria))
        ws.cell(row=row, column=SC(6), value=div_f(SL(5), row, SL(2), row))
        ws.cell(row=row, column=SC(6)).number_format = PCT_FMT
        _apply_numfmt(ws, row, row, (2, 3, 4, 5))
        _style_row(ws, row, len(headers))
        row += 1
    return row + 1


def write_summary_sheet(ws: Worksheet, agg_result: AggregationResult, ctx: SheetContext,
                         borrower_df: pd.DataFrame, config: dict) -> None:
    t = agg_result.tables
    row = 1

    row = _section_title(ws, row, "1. 총괄표")
    row = _section_overall_summary(ws, row, borrower_df, t["overall_summary"], config, ctx)

    row = _section_title(ws, row, "2. 차주형태별 분류")
    row = _section_borrower_type(ws, row, t["borrower_type"], ctx)

    row = _section_title(ws, row, "3. 금액별 분류")
    row = _section_amount_bucket(ws, row, t["amount_bucket"], ctx)

    row = _section_title(ws, row, "4. 담보물 종류별 분석")
    row = _section_property_type(ws, row, t["property_type"], ctx)

    row = _section_title(ws, row, "5. 담보물 지역별 분류")
    row = _section_region(ws, row, t["region"], ctx)
    row = _section_matrix(ws, row, t["region_x_property"], ctx)

    row = _section_title(ws, row, "6. 경매절차에 따른 분류")
    row = _section_foreclosure(ws, row, t["foreclosure"], ctx)

    row = _section_title(ws, row, "7. 회생절차에 따른 분류")
    row = _section_rehab(ws, row, t["rehab"], ctx)

    row = _section_title(ws, row, "8. 신용보증서 유무에 따른 분류")
    pools = sorted(borrower_df["pool_type"].dropna().unique())
    row = _section_guarantee(ws, row, pools, ctx)

    ws.column_dimensions["A"].width = 16  # holds the gray (그룹명) annotation only
    ws.column_dimensions[get_column_letter(SC(1))].width = 22
    for c in range(2, 8):
        ws.column_dimensions[get_column_letter(SC(c))].width = 15
    ws.freeze_panes = "A1"

    if agg_result.unclassified_property_types or agg_result.unclassified_regions:
        row += 1
        ws.cell(row=row, column=SC(1), value="※ classification_config.json에 분류 규칙이 없어 '기타'로 처리된 항목").font = NOTE_FONT
        row += 1
        if agg_result.unclassified_property_types:
            ws.cell(row=row, column=SC(1),
                    value="담보물 종류: " + ", ".join(sorted(agg_result.unclassified_property_types))).font = NOTE_FONT
            row += 1
        if agg_result.unclassified_regions:
            ws.cell(row=row, column=SC(1),
                    value="지역: " + ", ".join(sorted(agg_result.unclassified_regions))).font = NOTE_FONT


# ---------------------------------------------------------------------------

def write_report(frames: Dict[str, pd.DataFrame], agg_result: AggregationResult, config: dict,
                  output_path: str, program_name: str = "NPL Data Disk 정리 결과") -> None:
    wb = Workbook()
    wb.remove(wb.active)

    borrower_sheet, borrower_cols = write_borrower_sheet(wb, frames["borrower"], config, program_name)
    collateral_sheet, collateral_cols = write_collateral_sheet(
        wb, frames["collateral"], config,
        agg_result.unclassified_property_types, agg_result.unclassified_regions,
        borrower_sheet, borrower_cols, program_name)
    guarantee_sheet, guarantee_cols = write_guarantee_sheet(
        wb, frames["guarantee"], borrower_sheet, borrower_cols, program_name)
    rehab_sheet, rehab_cols = write_rehab_sheet(
        wb, frames["rehab"], borrower_sheet, borrower_cols, program_name)

    ctx = SheetContext(
        borrower_sheet=borrower_sheet, borrower_cols=borrower_cols,
        collateral_sheet=collateral_sheet, collateral_cols=collateral_cols,
        guarantee_sheet=guarantee_sheet, guarantee_cols=guarantee_cols,
        rehab_sheet=rehab_sheet, rehab_cols=rehab_cols,
    )

    summary_ws = wb.create_sheet(OUTPUT_SHEET_NAMES["summary"])
    write_summary_sheet(summary_ws, agg_result, ctx, frames["borrower"], config)

    wb.move_sheet(OUTPUT_SHEET_NAMES["summary"], offset=-(len(wb.sheetnames) - 1))
    wb.save(output_path)
